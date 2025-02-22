import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import networkx as nx
import random
import time
from datetime import datetime
import json
import os
import re
import openpyxl
import shutil


def f_net_newfolder():
    global TIME_ID, START_TIME
    global CODE_DIR, DATASETS_DIR, DEST_DIR, RUN_NUMBER
    
    print("\n-----------------------------------------------")
    TIME_ID = f'{datetime.now().strftime("%y%m%d-%H%M%S")}'
    START_TIME = f'{datetime.now().strftime("%H:%M:%S")}'
    print(f"\nRun id: {TIME_ID}")
    print(f"\nStart time: {START_TIME}")
    print("\n-----------------------------------------------")

    # Adding \.. to the end of the path. so abspath is the parrent folder

    # D:\Academic\Research\BrainSyncMaster\010-Networks\Code
    CODE_DIR = os.path.dirname(os.path.realpath(__file__))
    # D:\Academic\Research\BrainSyncMaster\010-Networks
    PARENT_DIR = os.path.abspath(CODE_DIR+"\\..")
    # D:\Academic\Research\BrainSyncMaster\
    PROJECT_DIR = os.path.abspath(CODE_DIR+"\\..\\..")
    # D:\Academic\Research\BrainSyncMaster\010-Networks\DATASETS\
    DATASETS_DIR = PARENT_DIR + "\\DATASETS\\"
    # D:\Academic\Research\BrainSyncMaster\Outputs\
    dest_dir = PROJECT_DIR + "\\Outputs\\"

    # Creat new folder (How many runs have been ran)
    L_pre_runs = list(filter(re.compile("Net ").match, os.listdir(dest_dir)))

    if len(L_pre_runs) != 0:
        # finding max run number
        RUN_NUMBER = (int((re.findall(" [0-9]{3} ", L_pre_runs[-1]))[0])) + 1
        # naming the folder
        if RUN_NUMBER < 10:
            RUN_NUMBER = f'Net 00{RUN_NUMBER}'
        elif RUN_NUMBER < 100:
            RUN_NUMBER = f'Net 0{RUN_NUMBER}'
        else:
            RUN_NUMBER = f'Net {RUN_NUMBER}'
    # if this is the first run
    else:
        RUN_NUMBER = 1
        RUN_NUMBER = f'Net 00{RUN_NUMBER}'


    DEST_DIR = os.path.join(f'{dest_dir}', f"{RUN_NUMBER} (id = {TIME_ID})")
    os.makedirs(DEST_DIR)   

    # To copy this code and json file destination folder
    shutil.copytree(PARENT_DIR + "\\Code" , DEST_DIR+"\\Initial Parameters", dirs_exist_ok=True)



def f_read_initial_data():
    global DICT_RUN, DICT_DATASETS, DICT_NETWORK
    
    with open(CODE_DIR+"\\Initial Networks Parameters.json", "r") as json_file:
        J_DATA      = json.load(json_file)
        
    DICT_RUN        = J_DATA["RUN"]
    DICT_DATASETS   = J_DATA["DATASETS"]
    DICT_NETWORK    = J_DATA["NETWORKS"]




def f_graph_weight(w_type, w_const, w_min, w_max, num_edges):
    
    arr_weight = np.zeros(num_edges)
    
    if(w_type == "NORMAL_RANDOM"):
        mu = (w_max + w_min) / 2
        sigma = (w_max - w_min) / 3
        arr_weight = np.random.normal(mu, sigma, num_edges)
        arr_weight[(arr_weight < w_min )|(arr_weight > w_max)] = mu

    elif(w_type == "UNIF_RANDOM"):
        arr_weight = np.random.uniform(low = w_min, high = w_max, size = num_edges)

    elif(w_type == "CONSTANT"):
        arr_weight = arr_weight + w_const
        print(arr_weight)

    arr_weight = np.array(list(map(int, arr_weight)))

    return arr_weight


def f_create_nodes_cent_atr_xlsx (new_xlsx_path, NUM_REP, THE_GRAPH, g_dircted):
    
    #to add new sheet to xlsx we create new data frame
    df_atributes = pd.DataFrame()
    df_atributes["Id"] = pd.DataFrame(sorted(list(THE_GRAPH.nodes)))
    df_atributes["Label"] = pd.DataFrame(sorted(list(THE_GRAPH.nodes)))
    df_atributes["DegreeCent"] = pd.DataFrame(sorted(nx.degree_centrality(THE_GRAPH).items()))[[1]]
    df_atributes["ClosenessCent"] = pd.DataFrame(sorted(nx.closeness_centrality(THE_GRAPH).items()))[[1]]
    df_atributes["BetweennessCent"] = pd.DataFrame(sorted(nx.betweenness_centrality(THE_GRAPH).items()))[[1]]
    df_atributes["EigenvectorCent"] = pd.DataFrame(sorted(nx.eigenvector_centrality_numpy(THE_GRAPH).items()))[[1]]
    
    if (g_dircted == "Undir"):
        sh_name = f"Nodes_Undir_Rep{str(NUM_REP+1)}"
        
    elif (g_dircted == "Dir"):
        sh_name = f"Nodes_Dir_Rep{str(NUM_REP+1)}"
        
    elif (g_dircted == "DataSet"):
        sh_name = "Nodes_Info"
        return df_atributes
        
    with pd.ExcelWriter(new_xlsx_path, mode = "a", engine = "openpyxl") as writer:
        df_atributes.to_excel(writer, sheet_name= sh_name, index = False)
        
        







def f_creat_networks():
    

    num_networks = DICT_NETWORK["NUM_TOPOLOGY"]
    for n in range(num_networks):
        
        
        g_name      = DICT_NETWORK['L_NAME_NET'][n]
        t_name      = DICT_NETWORK['L_NAME_TOPOLOGY'][n]
        g_nodes     = DICT_NETWORK['L_NUM_NODES'][n]
        g_ave_deg   = DICT_NETWORK['L_NUM_AVE_DEG'][n]
        g_p_rw      = DICT_NETWORK['L_P_REW'][n]
        g_dircted   = DICT_NETWORK['L_BOOL_DI'][n]
        g_w_type    = DICT_NETWORK['L_WEIGHT'][n]
        g_w_const   = DICT_NETWORK['L_CONSTANT_WEIGHT'][n]
        g_w_min     = DICT_NETWORK['L_MIN_WEIGHT'][n]
        g_w_max     = DICT_NETWORK['L_MAX_WEIGHT'][n]

        df_network_info = pd.DataFrame()

        df_network_info["Repeat"]       = [DICT_NETWORK["NUM_REP"]]
        df_network_info["Network"]      = [g_name]
        df_network_info["Directed"]     = [g_dircted]
        df_network_info["Nodes"]        = [g_nodes]
        df_network_info["AveDegree"]    = [g_ave_deg]
        df_network_info["RewiringProb"] = [g_p_rw]
        df_network_info["ConstWeight"]  = [g_w_const]
        df_network_info["MinWeight"]    = [g_w_min]
        df_network_info["MaxWeight"]    = [g_w_max]
        df_network_info["WeightType"]   = [g_w_type]
    
        df_network_info = df_network_info.T
        df_network_info = df_network_info.reset_index()
        df_network_info.columns = ["Network_Parameters", "Network_Values"] 
        
        # print(df_network_info)

        # To creat an empty xlsx file
        # new_xlsx_path = DEST_DIR + f"\\T{n}_{t_name}_N={g_nodes}_K={g_ave_deg}_P={g_p_rw}_D={g_dircted}_(id = {TIME_ID}).xlsx"
        new_xlsx_path = DEST_DIR + f"\\Net-{n+1}-{t_name}-{TIME_ID}.xlsx"
        wb = openpyxl.Workbook()
        wb.save(new_xlsx_path)
        
        sh_name = "Network_Info"
        with pd.ExcelWriter(new_xlsx_path, mode = "a", engine = "openpyxl") as writer:
            df_network_info.to_excel(writer, sheet_name= sh_name, index = False)
        
        print(f"\n\tTOPOLOGY {n+1} of {num_networks}")

        for r in range(DICT_NETWORK["NUM_REP"]):
            
            print(f"\t{g_name}, REAPEAT: {r+1} is done!")
            
            if (g_name == "ER_Graph"):
                #plt_title = "Erdos-Renyi Graph"
                ER_p = g_ave_deg / (g_nodes - 1)
                THE_GRAPH = nx.erdos_renyi_graph(g_nodes, ER_p)

            elif (g_name == "BA_Graph"):
                #plt_title = "Barabasi-Albert Graph"
                BA_m = int(g_ave_deg / 2)
                THE_GRAPH = nx.barabasi_albert_graph(g_nodes, BA_m)
            
            elif (g_name == "WS_Graph"):
                #plt_title = "Watts-Strogatz Graph" 
                THE_GRAPH = nx.watts_strogatz_graph(g_nodes, k = g_ave_deg, p = g_p_rw)  
            
            elif (g_name == "Star_Graph"):
                #plt_title = "Star Graph"
                THE_GRAPH = nx.star_graph(g_nodes) 
    
            df_graph_edgelist = nx.to_pandas_edgelist(THE_GRAPH).copy()
            df_graph_edgelist = df_graph_edgelist.sort_values(by=['source', 'target'])

            
            df_graph_edgelist["weight"] = f_graph_weight(g_w_type, g_w_const, g_w_min, g_w_max, df_graph_edgelist.shape[0])

            
            if (g_dircted == "False" or g_dircted == "Both"):
                df_undir_edgelist = df_graph_edgelist.copy()
                
                df_undir_edgelist = df_undir_edgelist.sort_values(by=['source', 'target'])
                
                
                # set new edge list to the graph
                THE_GRAPH = nx.from_pandas_edgelist(df_undir_edgelist)
                
                f_create_nodes_cent_atr_xlsx(new_xlsx_path, r, THE_GRAPH, "Undir")
                
                temp_df = pd.DataFrame()
                temp_df['source'] = df_undir_edgelist['target'].copy()
                temp_df['target'] = df_undir_edgelist['source'].copy()
                temp_df['weight'] = df_undir_edgelist['weight'].copy()
                
                df_undir_edgelist = pd.concat([df_undir_edgelist, temp_df])
                df_undir_edgelist.reset_index(drop=True, inplace=True)
                df_undir_edgelist = df_undir_edgelist.sort_values(by=['source', 'target'])
                
                sh_name = f"Edge_Undir_Rep{str(r+1)}"
                with pd.ExcelWriter(new_xlsx_path, mode = "a", engine = "openpyxl") as writer:
                    df_undir_edgelist.to_excel(writer, sheet_name= sh_name, index = False)
                
        
            if (g_dircted == "True" or g_dircted == "Both"):
                
                e_edgelist_df = df_graph_edgelist[0: :2].copy()
                o_edgelist_df = df_graph_edgelist[1: :2].copy()

                #change order of columns
                e_edgelist_df = e_edgelist_df[['target','source','weight']]
                e_edgelist_df.columns = ['source','target','weight']
                
                df_dir_edgelist = pd.concat([e_edgelist_df, o_edgelist_df])
                df_dir_edgelist = df_dir_edgelist.sort_values(by=['source', 'target'])
                
                df_dir_edgelist.reset_index(drop=True, inplace=True)
             
                THE_DIR_GRAPH = nx.from_pandas_edgelist(df_dir_edgelist)
                
                f_create_nodes_cent_atr_xlsx(new_xlsx_path, r, THE_DIR_GRAPH, "Dir")
                
                sh_name = f"Edge_Dir_Rep{str(r+1)}"
                with pd.ExcelWriter(new_xlsx_path, mode = "a", engine = "openpyxl") as writer:
                    df_dir_edgelist.to_excel(writer, sheet_name= sh_name, index = False) 
                    
                    
                
                    
            # print(df_undir_edgelist)
            # input() 
            
        wb=openpyxl.load_workbook(new_xlsx_path)
        wb.remove(wb['Sheet'])
        wb.save(new_xlsx_path)
    


def f_read_dataset():

    global L_address_input_DataSets
    #make a list of all xlsx files' addresses from the "Networks" folder
    L_address_input_DataSets = list(filter(re.compile(".*xlsx").match, os.listdir(DATASETS_DIR)))
    
    for ds in range (len(L_address_input_DataSets)):
        ds_address = L_address_input_DataSets[ds]
        DF_Network = pd.read_excel(DATASETS_DIR + ds_address, sheet_name = "Network_Info")
        DF_Nodes = pd.read_excel(DATASETS_DIR + ds_address, sheet_name = "Nodes_Info")
        DF_Edges = pd.read_excel(DATASETS_DIR + ds_address, sheet_name = "Edges_Info")

        # set new edge list to the graph
        THE_GRAPH = nx.from_pandas_edgelist(DF_Edges)
        
        df_nods_attr = f_create_nodes_cent_atr_xlsx("no_addres", "no_repeat", THE_GRAPH, "DataSet")
        
        DF_Nodes["DegreeCent"] = df_nods_attr["DegreeCent"]
        new_name = ds_address.split('.xlsx')[0]
        new_xlsx_path = DEST_DIR + f"\\Net-{ds+1}-{new_name}-{TIME_ID}.xlsx"
        wb = openpyxl.Workbook()
        wb.save(new_xlsx_path)

        with pd.ExcelWriter(new_xlsx_path, mode = "a", engine = "openpyxl") as writer:
            DF_Network.to_excel(writer, sheet_name= "Network_Info", index = False)
            DF_Nodes.to_excel(writer, sheet_name= "Nodes_Info", index = False)
            DF_Edges.to_excel(writer, sheet_name= "Edges_Info", index = False)
            
        wb=openpyxl.load_workbook(new_xlsx_path)
        wb.remove(wb['Sheet'])
        wb.save(new_xlsx_path)



def f_main():
    
    global END_TIME
    
    f_net_newfolder()
    f_read_initial_data()
    
    if(DICT_RUN["READ_FROM"] == "NETWORKS"):
        f_creat_networks()

    elif(DICT_RUN["READ_FROM"] == "DATASETS"):
        f_read_dataset()
        
    print("\n-----------------------------------------------")
    print(f"\nRun id: {TIME_ID}")

    END_TIME = f'{datetime.now().strftime("%H:%M:%S")}'
    print(f"\nStart time: {START_TIME}")
    print(f"End time: {END_TIME}")
    print("\n-----------------------------------------------")


f_main()

